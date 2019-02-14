from django.core.files.storage import FileSystemStorage
from django.views.decorators.csrf import csrf_exempt
from rest_framework.renderers import JSONRenderer
from django.core.validators import validate_email
from django.shortcuts import render, redirect
from rest_framework.parsers import JSONParser
from passlib.hash import pbkdf2_sha256
from django.http import HttpResponse
from django.conf import settings
from mailjet_rest import Client
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from user.serializers import *
from user.models import *
from datetime import datetime, timedelta
from django.utils import timezone
import requests
import random
import json
import jwt

import os

import firebase_admin
from firebase_admin import credentials, messaging



# Create your views here.
#
# CodeAccount(code="PFO4819380", image="users/CONFIEP.png").save()
# CodeAccount(code="PFO9294795", image="users/SARCC.png").save()
# CodeAccount(code="PFO9833938", image="users/PERÚCAMARAS.png").save()
# CodeAccount(code="PFO7461508", image="users/SNI.png").save()
# CodeAccount(code="PFO9635980", image="users/SNMPE.png").save()
# CodeAccount(code="PFO1293902", image="users/SNP.png").save()
# CodeAccount(code="PFO8015687", image="users/APESEG.png").save()
# CodeAccount(code="PFO5407593", image="users/CANATUR.png").save()
# CodeAccount(code="PFO8709620", image="users/CAPECO.png").save()
# CodeAccount(code="PFO2270144", image="users/CCL.png").save()
# CodeAccount(code="PFO5133451", image="users/ASBANC.png").save()
# CodeAccount(code="AUT1729571", image="").save()
# CodeAccount(code="AUT9374612", image="").save()
# CodeAccount(code="AUT1471237", image="").save()
# CodeAccount(code="AUT1927316", image="").save()
# CodeAccount(code="AUT9237412", image="").save()
# CodeAccount(code="AUT6348341", image="").save()
# CodeAccount(code="AUT8246192", image="").save()
# CodeAccount(code="AUT2715718", image="").save()
# CodeAccount(code="AUT8746192", image="").save()
# CodeAccount(code="AUT6410273", image="").save()
# CodeAccount(code="AUT7162381", image="").save()
# CodeAccount(code="AUT2187213", image="").save()
# CodeAccount(code="AUT3450127", image="").save()
# CodeAccount(code="AUT8263171", image="").save()
# CodeAccount(code="AUT6481293", image="").save()
# CodeAccount(code="AUT9162361", image="").save()
# CodeAccount(code="AUT6182323", image="").save()
# CodeAccount(code="AUT9252891", image="").save()
# CodeAccount(code="AUT6182868", image="").save()
# CodeAccount(code="AUT6517823", image="").save()
# CodeAccount(code="AUT9762612", image="").save()
# CodeAccount(code="AUT7615283", image="").save()
# CodeAccount(code="AUT8616232", image="").save()
# CodeAccount(code="AUT6528915", image="").save()
# CodeAccount(code="AUT5288255", image="").save()

#-------------------------------GLOBAL VARIABLES-------------------------------
cred = credentials.Certificate("gaed-812dc-firebase-adminsdk-etqv6-c5c695567a.json")
firebase_admin.initialize_app(cred)

#-------------------------------UTILS-------------------------------

class JSONResponse(HttpResponse):

    def __init__(self, data, **kwargs):
        content = JSONRenderer().render(data)
        kwargs['content_type'] = 'application/json; charset=utf-8'
        super(JSONResponse, self).__init__(content,**kwargs)

def consoleLog(text='Información',data= ''):
    print('########## {text} : {data} ##########'.format(text=text, data=data))


def generate_identifier():
    identifier = str(random.randint(1000000000,9999999999))
    try:
        Agent.objects.get(identifier = identifier)
        return generate_identifier(identifier)
    except Agent.DoesNotExist as e:
        return identifier

def generate_token(user):
    payload = {
        "phone": user['phone'],
        "identifier": user['identifier'],
        "authority": False
    }
    encoded = jwt.encode(payload, settings.JWT_SECRET_KEY, algorithm='HS256')
    return encoded.decode("utf-8")

def get_data_from_request(request):
    return request.POST.dict() if len(request.POST) else JSONParser().parse(request)


#--------------------------------------------- AGENT ---------------------------------------------

@csrf_exempt
def register(request):
    if request.method == 'POST':
        user = get_data_from_request(request)
        user['password'] = pbkdf2_sha256.encrypt(user['password'],rounds=12000,salt_size=32)
        user['identifier'] = generate_identifier()
        user['token'] = generate_token(user)
        prefix = user['code'][0:3]
        if prefix == "AUT":
            print("AUT")
            try:
                code = CodeAccount.objects.get(code=user['code'])
                if code.used:
                    return JSONResponse({'error':{'code': 401, 'message':'Código usado'}})
                else:
                    serializer = AgentSerializer(data=user)
            except Exception as e:
                return JSONResponse({'error':{'code': 401, 'message':'Código incorrecto'}})
        elif prefix == "PFO":
            print("PFO")
            try:
                code = CodeAccount.objects.get(code=user['code'])
                if code.used:
                    return JSONResponse({'error':{'code': 401, 'message':'Código usado'}})
                else:
                    user["photo_url"] = code.image
                    serializer = EmpresaFocalSerializer(data=user)
            except Exception as e:
                return JSONResponse({'error':{'code': 401, 'message':'Código incorrecto'}})

        else:
            return JSONResponse({'error':{'code': 401, 'message':'Código incorrecto'}})

        if serializer.is_valid():
            try:
                validate_email(user['email'])
            except ValidationError as e:
                return JSONResponse({'error':{'code': 401, 'message':'Correo electrónico con formato incorrecto'}})
            serializer.save()
            code.used = True
            code.save()
            #send_email_register(user)
            return JSONResponse({})
        if "phone" in serializer.errors:
            return JSONResponse({'error':{'code': 401, 'message':'Este teléfono ya ha sido usado'}})
        else:
            return JSONResponse({'error':{'code': 401, 'message':'Este correo ya ha sido usado'}})
    return JSONResponse({'error':{'code': 405, 'message':'Método Http inválido'}})

@csrf_exempt
def login(request):
    print("Login")

    if request.method == 'POST':
        credentials = get_data_from_request(request)
        try:
            user = Agent.login(credentials)
            token = user.token
            print("USER", token)
            return JSONResponse({'token':token})
        except Agent.DoesNotExist as e:
            print(e)

        try:
            focal = EmpresaFocal.login(credentials)
            token = focal.token
            print("FOCAL", token)
            return JSONResponse({'token':token})
        except EmpresaFocal.DoesNotExist as e:
            print(e)

        return JSONResponse({'error':{'code': 401, 'message':'Correo o contraseña incorrecto'}})
    return JSONResponse({'error':{'code': 405, 'message':'Método Http inválido'}})

@csrf_exempt
def user_info(request):
    if request.method == 'GET':
        if 'HTTP_AUTHORIZATION' not in request.META: return JSONResponse({'error':{'code': 403, 'message':'Error de autorización'}}, status=403)
        token = request.META['HTTP_AUTHORIZATION'].split(" ")[1]
        try:
            agent = Agent.objects.get(token=token)
            print("USER", agent.name)
            return JSONResponse({'login':{'agent':agent.as_dict_agent()}})
        except Exception as e:
            print(e)

        try:
            focal = EmpresaFocal.objects.get(token=token)
            print("FOCAL", focal.name)
            print(focal.as_dict_agent())
            return JSONResponse({'login':{'focal':focal.as_dict_agent()}})
        except Exception as e:
            print(e)

        return JSONResponse({'error':{'code': 401, 'message':'Correo o contraseña incorrecto'}})
    else:
        return JSONResponse({'error':{'code': 405, 'message':'Método Http inválido'}})



#---------------------------------SOLICITUDE-------------------------------------
@csrf_exempt
def donate(request, identifier):
    """
    List all code request, or create a new request.
    """
    try:
        user = Agent.objects.get(identifier=identifier)
    except Agent.DoesNotExist:
        return HttpResponse(status=404)
    if request.method == 'POST':
        token = request.META['HTTP_AUTHORIZATION'].split(" ")[1]

        if token == user.token:
            data = get_data_from_request(request)
            return JSONResponse({}, status=201)

        #     print(data)
        #     data['agent'] = user.id
        #     # data['date'] = "2018-10-30T04:05"
        #     # data['deadline'] = "2018-11-1T04:05"
        #     serializer = SolicitudeSerializer(data=data)
        #     if serializer.is_valid():
        #         serializer.save()
        #         solicitude = Solicitude.objects.get(pk=serializer.data['id'])
        #         if 'items' in data:
        #             for item in json.loads(request.POST.dict()['items']):
        #                 Item(solicitude= solicitude, product= item['product'], amount= item['amount']).save()
        #         if 'product_list' in data:
        #             for item in json.loads(request.POST.dict()['product_list']):
        #                 Item(solicitude= solicitude, product= item['product'], amount= item['amount']).save()
        #         solicitudes = Solicitude.objects.filter(closed=False, accepted=True, deadline__gte=datetime.now()).order_by('-id')
        #         solicitudes_dict = [solicitude.as_dict_agent() for solicitude in solicitudes]
        #         return JSONResponse({'solicitudes':solicitudes_dict}, status=201)
        #     print(serializer.errors)
            return JSONResponse(serializer.errors, status=400)
        else:
            return HttpResponse(status=401)

    else:
        return HttpResponse(status=405)

#---------------------------------SOLICITUDE-------------------------------------
@csrf_exempt
def solicitude_list(request, identifier):
    """
    List all code request, or create a new request.
    """

    if request.method == 'GET':
        solicitudes = Solicitude.objects.filter(closed=False, accepted=True, deadline__gte=datetime.now()).order_by('-id')
        solicitudes_dict = [solicitude.as_dict_agent() for solicitude in solicitudes]
        return JSONResponse({'solicitudes':solicitudes_dict})

    elif request.method == 'POST':
        try:
            user = Agent.objects.get(identifier=identifier)
        except Agent.DoesNotExist:
            return JSONResponse({'error':{'code': 403, 'message':'Error de autorización'}}, status=403)
        token = request.META['HTTP_AUTHORIZATION'].split(" ")[1]

        if token == user.token:
            data = get_data_from_request(request)
            data['agent'] = user.id
            # data['date'] = "2018-10-30T04:05"
            # data['deadline'] = "2018-11-1T04:05"
            serializer = SolicitudeSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                solicitude = Solicitude.objects.get(pk=serializer.data['id'])
                if 'product_list' in data:
                    for item in json.loads(request.POST.dict()['product_list']):
                        Item(solicitude= solicitude, product= item['product'], total= item['total'], remaining= item['remaining']).save()
                solicitudes = Solicitude.objects.filter(closed=False, accepted=True, deadline__gte=datetime.now()).order_by('-id')
                solicitudes_dict = [solicitude.as_dict_agent() for solicitude in solicitudes]

                #send_email("SOLICITUD CREADA", solicitude)
                #send_notification("Nueva solicitud {} creada".format(solicitude.title), solicitude)
                print({'solicitudes':solicitudes_dict})
                return JSONResponse({'solicitudes':solicitudes_dict}, status=201)
            print(serializer.errors)
            return JSONResponse({'error':{'code': 401, 'message':'Datos no válidos '}})
        else:
            return JSONResponse({'error':{'code': 401, 'message':'Correo o contraseña incorrecto'}})
    else:
        return JSONResponse({'error':{'code': 405, 'message':'Método Http inválido'}})

from django.http import QueryDict

@csrf_exempt
def solicitude_detail(request, identifier, pk_solicitude):
    """
    Retrieve, update or delete a request.
    """
    try:
        solicitude = Solicitude.objects.get(pk=pk_solicitude)
    except Agent.DoesNotExist:
        return HttpResponse(status=404)
    except Solicitude.DoesNotExist:
        return HttpResponse(status=404)

    if request.method == 'GET':
        #serializer = SolicitudeSerializer(solicitude)
        return JSONResponse({'solicitude':solicitude.as_dict_agent()})
    elif request.method == 'DELETE':
        token = request.META['HTTP_AUTHORIZATION'].split(" ")[1]
        solicitude.change_to_closed()
        solicitudes = user.solicitude_set.all()
        solicitudes_dict = [solicitude.as_dict_agent() for solicitude in solicitudes]
        # photos = job_request.photo_set.all()
        # for photo in photos:
        #     simple_delete_job_request(photo)
        # job_request.delete()
        return JSONResponse(solicitudes_dict)
    # elif request.method == 'PUT':
    #     data = QueryDict(request.body)
    #     token = request.META['HTTP_AUTHORIZATION'].split(" ")[1]
    #     solicitude = Solicitude.objects.filter(pk=pk_solicitude).first()
    #     solicitude_temp = Solicitude.objects.get(pk=solicitude.pk)
    #
    #     focal = Focal(name=data['name'], RUC_or_DNI=data['ruc_or_dni'])
    #     focal.save()
    #     solicitude.focal = focal
    #     solicitude.closed = True
    #     solicitude.save()
    #     solicitude.focal = None
    #     solicitude.pk = None
    #     solicitude.closed = False
    #     solicitude.save()
    #
    #     data = QueryDict(request.body)
    #
    #     amount_list = json.loads(data['product_list'])
    #     for (index, item) in enumerate(Solicitude.objects.get(pk=solicitude_temp.pk).item_set.all()):
    #         print(index)
    #         item.help = amount_list[index]
    #         item.save()
    #         item.pk = None
    #         item.amount = item.amount - amount_list[index]
    #         if item.amount != 0:
    #             delete_item = False
    #             item.help = 0
    #             item.solicitude = solicitude
    #             item.save()
    #
    #     if delete_item:
    #         solicitude.delete()
    #     send_email("DONACIÓN REALIZADA", solicitude)
    #     return JSONResponse({}, status=200)
    elif request.method == 'PUT':
        solicitude = Solicitude.objects.filter(pk=pk_solicitude).first()
        empresa_focal = EmpresaFocal.objects.filter(identifier=identifier).first()

        if solicitude.closed:
            return JSONResponse({'error':{'code': 401, 'message':'La solicitud ya ha sido cerrada'}})

        data = QueryDict(request.body)
        help = Help(name=data['name'], RUC_or_DNI=data['ruc_or_dni'], empresa_focal=empresa_focal, solicitude=solicitude)
        help.save()

        amount_list = json.loads(data['product_list'])
        for (index, item) in enumerate(solicitude.item_set.all().exclude(remaining=0)):
            HelpItem(help= help, item= item, amount= amount_list[index]).save()
            item.remaining = item.remaining - amount_list[index]
            item.save()

        if solicitude.item_set.all().exclude(remaining=0).count() == 0:
            solicitude.change_to_closed()
        #send_email("DONACIÓN REALIZADA", solicitude)
        return JSONResponse({}, status=200)
    else:
        return JSONResponse({'error':{'code': 405, 'message':'Método Http inválido'}})

@csrf_exempt
def update_item(request, identifier, pk_solicitude, pk_item):
    if request.method == 'PUT':
        solicitude = Solicitude.objects.filter(pk=pk_solicitude).first()
        empresa_focal = EmpresaFocal.objects.filter(identifier=identifier).first()
        item = solicitude.item_set.get(pk=pk_item)
        if solicitude.closed:
            return JSONResponse({'error':{'code': 401, 'message':'La solicitud ya ha sido cerrada'}})

        data = QueryDict(request.body)
        help = Help(name=data['name'], RUC_or_DNI=data['ruc_or_dni'], empresa_focal=empresa_focal, item=item, amount= int(data['amount']))
        help.save()

        if item.remaining <= 0:
            return JSONResponse({'error':{'code': 401, 'message':'El requerimiento ya ha sido atendido'}})
        else:
            item.remaining = item.remaining - int(data['amount'])
            item.save()

        if solicitude.item_set.all().exclude(remaining=0).count() == 0:
            solicitude.change_to_closed()
            send_email_closed('Hola {}, tu solicitud "{}" ha sido cubierta por completo.'.format(solicitude.agent.name, solicitude.title), solicitude.agent )
        #send_email("DONACIÓN REALIZADA", solicitude)
        return JSONResponse({"solicitude": solicitude.as_dict_agent()})
    else:
        return JSONResponse({'error':{'code': 405, 'message':'Método Http inválido'}})


@csrf_exempt
def upload_image(request):
    if request.method == 'POST' and request.FILES['upload']:
        image = request.FILES['upload']
        fs = FileSystemStorage()
        path = str(random.randint(0,9999999999)) + ".png"
        filename = fs.save(path, image)
        return JSONResponse({'image_url': filename}, status=201)
    return JSONResponse({'error':{'code': 405, 'message':'Método Http inválido'}})

#---------------------------------EXCEL-------------------------------------
from openpyxl.styles import Font

@csrf_exempt
def export_excel(request):
    if request.method == 'GET':
        closed = request.GET["closed"]
        period = request.GET["period"]

        current = {
            "week":datetime.today().weekday(),
            "month":datetime.today().day,
            "all": 9999
        }

        day_offset = datetime.now().date() - timedelta(days=current[period])

        if closed == "all":
            solicitudes = Solicitude.objects.filter(date__gte=day_offset).order_by('-date')
        elif closed == "closed":
            solicitudes = Solicitude.objects.filter(closed=True, date__gte=day_offset).order_by('-date')
        elif closed == "open":
            solicitudes = Solicitude.objects.filter(closed=False, date__gte=day_offset).order_by('-date')

        wb = Workbook()
        ws = wb.active
        ws2 = wb.create_sheet()
        ws3 = wb.create_sheet()

        ws.title = 'Solicitudes'
        ws2.title = 'Requerimientos'
        ws3.title = 'Ayudas'

        ws.merge_cells('B2:D2')
        ws2.merge_cells('B2:D2')
        ws3.merge_cells('B2:D2')

        redFill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
        headerFill = PatternFill(start_color='E2434B', end_color='E2434B', fill_type='solid')
        bodyFill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        titleFill = PatternFill(start_color='CA182F', end_color='CA182F', fill_type='solid')

        headerTexts = ["Código","Autor","Título","Emergencia","Región","Provincia","Distrito","Prioridad","Fecha","Cerrado","Nombre","DNI","Teléfono"]
        headerTexts2 = ["Código de Solicitud","Código de Requerimiento","Requerimiento","Total","Faltante"]
        headerTexts3 = ["Código de Requerimiento","Código de Ayuda","Institución","Representante","RUC o DNI","Cantidad"]

        ws['B2'] = "LISTA DE SOLICITUDES"
        ws2['B2'] = "LISTA DE REQUERIMIENTOS"
        ws3['B2'] = "LISTA DE AYUDAS"

        ws.cell(row=2, column=2).fill = titleFill
        ws.cell(row=2, column=2).font = Font(color="FFFFFF")

        ws2.cell(row=2, column=2).fill = titleFill
        ws2.cell(row=2, column=2).font = Font(color="FFFFFF")

        ws3.cell(row=2, column=2).fill = titleFill
        ws3.cell(row=2, column=2).font = Font(color="FFFFFF")

        for i in range(3, 16):
            cell = ws.cell(row=4, column=i)
            cell.value = headerTexts[i - 3]
            cell.fill = headerFill
            cell.font = Font(color="FFFFFF")

        ws.cell(row=3, column=13).value = "Receptor"
        ws.cell(row=3, column=13).fill = headerFill
        ws.cell(row=3, column=13).font = Font(color="FFFFFF")
        ws.merge_cells('M3:O3')

        for i in range(4, 9):
            cell = ws2.cell(row=4, column=i)
            cell.value = headerTexts2[i - 4]
            cell.fill = headerFill
            cell.font = Font(color="FFFFFF")

        for i in range(3, 9):
            cell = ws3.cell(row=4, column=i)
            cell.value = headerTexts3[i - 3]
            cell.fill = headerFill
            cell.font = Font(color="FFFFFF")

        start_index_1 = 5
        start_index_2 = 5
        start_index_3 = 5


        #SOLICITUDES LIST
        for solicitude in solicitudes:
            ws.cell(row=start_index_1, column=2).value = '=HYPERLINK("#Requerimientos!D{}:H{}","Ver Requerimientos")'.format(str(start_index_2), str(start_index_2 + solicitude.item_set.count() - 1))
            ws.cell(row=start_index_1, column=2).font = Font(color="E2434B")

            ws.cell(row=start_index_1, column=3).value = "#GEAD" + str(solicitude.pk)
            ws.cell(row=start_index_1, column=3).fill = bodyFill

            ws.cell(row=start_index_1, column=4).value = solicitude.agent.name
            ws.cell(row=start_index_1, column=4).fill = bodyFill

            ws.cell(row=start_index_1, column=5).value = solicitude.title
            ws.cell(row=start_index_1, column=5).fill = bodyFill

            ws.cell(row=start_index_1, column=6).value = solicitude.emergency
            ws.cell(row=start_index_1, column=6).fill = bodyFill

            ws.cell(row=start_index_1, column=7).value = solicitude.region
            ws.cell(row=start_index_1, column=7).fill = bodyFill

            ws.cell(row=start_index_1, column=8).value = solicitude.province
            ws.cell(row=start_index_1, column=8).fill = bodyFill

            ws.cell(row=start_index_1, column=9).value = solicitude.district
            ws.cell(row=start_index_1, column=9).fill = bodyFill

            ws.cell(row=start_index_1, column=10).value = solicitude.priority
            ws.cell(row=start_index_1, column=10).fill = bodyFill

            ws.cell(row=start_index_1, column=11).value = solicitude.date.strftime('%d/%m/%y')
            ws.cell(row=start_index_1, column=11).fill = bodyFill

            ws.cell(row=start_index_1, column=12).value = "Cerrado" if solicitude.closed else "Abierto"
            ws.cell(row=start_index_1, column=12).fill = bodyFill

            if solicitude.closed:
                ws.cell(row=start_index_1, column=12).fill = redFill

            ws.cell(row=start_index_1, column=13).value = solicitude.receiver_name
            ws.cell(row=start_index_1, column=13).fill = bodyFill

            ws.cell(row=start_index_1, column=14).value = solicitude.receiver_dni
            ws.cell(row=start_index_1, column=14).fill = bodyFill

            ws.cell(row=start_index_1, column=15).value = solicitude.receiver_phone
            ws.cell(row=start_index_1, column=15).fill = bodyFill

            start_index_1+=1

            #REQUERIMIENTOS LIST
            for item in solicitude.item_set.all():
                ws2.cell(row=start_index_2, column=2).value = '=HYPERLINK("#Solicitudes!C{}:O{}","Volver a la Solicitud")'.format(str(start_index_1 - 1), str(start_index_1 - 1))
                ws2.cell(row=start_index_2, column=2).font = Font(color="E2434B")

                ws2.cell(row=start_index_2, column=3).value = '=HYPERLINK("#Ayudas!C{}:H{}","Ver Ayudas")'.format(str(start_index_3), str(start_index_3 + item.help_set.count() - 1))
                ws2.cell(row=start_index_2, column=3).font = Font(color="E2434B")

                ws2.cell(row=start_index_2, column=4).value = "#GEAD" + str(solicitude.pk)
                ws2.cell(row=start_index_2, column=4).fill = bodyFill

                ws2.cell(row=start_index_2, column=5).value = item.pk
                ws2.cell(row=start_index_2, column=5).fill = bodyFill

                ws2.cell(row=start_index_2, column=6).value = item.product
                ws2.cell(row=start_index_2, column=6).fill = bodyFill

                ws2.cell(row=start_index_2, column=7).value = item.total
                ws2.cell(row=start_index_2, column=7).fill = bodyFill

                ws2.cell(row=start_index_2, column=8).value = item.remaining
                ws2.cell(row=start_index_2, column=8).fill = bodyFill

                start_index_2+=1

                #AYUDAS LIST
                for help in item.help_set.all():
                    ws3.cell(row=start_index_3, column=2).value = '=HYPERLINK("#Requerimientos!D{}:H{}","Volver a Requerimiento")'.format(str(start_index_2 - 1), str(start_index_2 - 1))
                    ws3.cell(row=start_index_3, column=2).font = Font(color="E2434B")

                    ws3.cell(row=start_index_3, column=3).value = item.pk
                    ws3.cell(row=start_index_3, column=3).fill = bodyFill

                    ws3.cell(row=start_index_3, column=4).value = help.pk
                    ws3.cell(row=start_index_3, column=4).fill = bodyFill

                    ws3.cell(row=start_index_3, column=5).value = help.empresa_focal.name
                    ws3.cell(row=start_index_3, column=5).fill = bodyFill

                    ws3.cell(row=start_index_3, column=6).value = help.name
                    ws3.cell(row=start_index_3, column=6).fill = bodyFill

                    ws3.cell(row=start_index_3, column=7).value = help.RUC_or_DNI
                    ws3.cell(row=start_index_3, column=7).fill = bodyFill

                    ws3.cell(row=start_index_3, column=8).value = help.amount
                    ws3.cell(row=start_index_3, column=8).fill = bodyFill

                    start_index_3+=1

        file_name = "ReporteSolicitudes {0}.xlsx".format(datetime.now().strftime("%d/%m/%y %H:%M"))
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(file_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response



#---------------------------------PASSWORD-------------------------------------
@csrf_exempt
def forgotten_password(request):
    if request.method == 'POST':
        data = get_data_from_request(request)
        try:
            agent = Agent.objects.get(email=data['email'])
            payload = {
                "email": agent.email,
                "exp": (timezone.now() + timezone.timedelta(hours=2)).timestamp()
            }
            _token = jwt.encode(payload, settings.JWT_SECRET_KEY, algorithm='HS256').decode("utf-8")
            send_forgotten_password(
            "Hola {}, dirígase al siguiente link para poder reestablecer su contraseña: \n {}".format(agent.name, settings.HOST + "users/reset_password/?token=" + _token), agent
            )
            return JSONResponse({},status=200)
        except Agent.DoesNotExist as e:
            print(e)

        try:
            focal = EmpresaFocal.objects.get(email=data['email'])
            payload = {
                "email": focal.email,
                "exp": (timezone.now() + timezone.timedelta(hours=2)).timestamp()
            }
            _token = jwt.encode(payload, settings.JWT_SECRET_KEY, algorithm='HS256').decode("utf-8")
            send_forgotten_password(
            "Hola {}, dirígase al siguiente link para poder reestablecer su contraseña: \n {}".format(focal.name, settings.HOST + "users/reset_password/?" + _token), focal
            )
            return JSONResponse({},status=200)
        except EmpresaFocal.DoesNotExist as e:
            print(e)
        return JSONResponse({'error':{'code': 401, 'message':'Este correo no está registrado'}})
    return JSONResponse({'error':{'code': 405, 'message':'Método Http inválido'}})

































#-------------------------------MAIL AND PUSH-------------------------------
API_KEY = '23e63458d588b10f67434ac7ca40b40e'
API_SECRET = '6108aa38fb2fa32124706e65af2b0c5c'
mailjet = Client(auth=(API_KEY, API_SECRET), version='v3.1')

data = {
  'Messages': [
                {
                        "From": {
                                "Email": "anthony.delpozo.m@gmail.com",
                                "Name": "Anthony Del Pozo"
                        },
                        "To": [
                                {
                                        "Email": "delan1997@gmail.com",
                                        "Name": "passenger 1"
                                }
                        ],
                        "TemplateID": 622922,
                        "TemplateLanguage": True,
                        "Subject": "Your email flight plan!",
                        "TextPart": "Dear passenger 1, welcome to Mailjet! May the delivery force be with you!",
                        "HTMLPart": "<h3 style='color:red; border: 1px solid blue; padding:10px; margin:20px'>Dear passenger 1, welcome to Mailjet!</h3><br />May the delivery force be with you!"
                }
        ]
}
# result = mailjet.send.create(data=data)
# print(result.status_code)
# print(result.__dict__)

def update_fcm_token(request):
    if request.method == 'POST':
        data = get_data_from_request(request)

        token = data["token"]
        new_fcm_token = data["new_fcm_token"]

        try:
            user = Agent.objects.get(token=token)
            user.fcm_token = new_fcm_token
            print("USER", new_fcm_token)
            return JSONResponse({'user':user})
        except Agent.DoesNotExist as e:
            print(e)

        try:
            focal = EmpresaFocal.objects.get(token=token)
            focal.fcm_token = new_fcm_token
            print("FOCAL", new_fcm_token)
            return JSONResponse({'user':focal})
        except EmpresaFocal.DoesNotExist as e:
            print(e)

        return JSONResponse({'error':{'code': 401, 'message':'Token incorrecto'}})
    return JSONResponse({'error':{'code': 405, 'message':'Método Http inválido'}})

def send_email_accepted(user, solicitude):
    pass
    email = {
        'FromName': 'GEAD APP',
        'FromEmail': 'anthony.delpozo.m@gmail.com',
        'Subject': "Solicitud aceptada",
        'Text-Part': "Hola {}, su solicitud '{}' ha sido aceptada.".format(user.name, solicitude.title),
        'Recipients': [{'Email': user.email}]
    }
    print(email)
    mailjet.send.create(email)

    # message = messaging.Message(
    #     data={
    #         "title":"Solicitud Aceptada",
    #         "body":"Hola {}, su solicitud '{}' ha sido aceptada.".format(user.name, solicitude.title)
    #     },
    #     token= user.fcm_token,
    # )
    # response = messaging.send(message)

def send_email_register(user):
    pass
    email = {
        'FromName': 'GEAD APP',
        'FromEmail': 'anthony.delpozo.m@gmail.com',
        'Subject': "Cuenta registrada",
        'Text-Part': "Hola {}, bienvenido a la familia GEAD, su cuenta ha sido creada exitosamente.".format(user['name']),
        'Recipients': [{'Email': user['email']}]
    }
    mailjet.send.create(email)

def send_email(message, solicitude):
    pass
    email = {
        'FromName': 'GEAD APP',
        'FromEmail': 'anthony.delpozo.m@gmail.com',
        'Subject': message,
        'Text-Part': message,
        'Recipients': [{'Email': 'delan1997@gmail.com'}]
    }
    mailjet.send.create(email)

def send_email_closed(message, agent):
    pass
    email = {
        'FromName': 'GEAD APP',
        'FromEmail': 'anthony.delpozo.m@gmail.com',
        'Subject': message,
        'Text-Part': message,
        'Recipients': [{'Email': agent.email}]
    }
    mailjet.send.create(email)

def send_forgotten_password(message, user):
    pass
    email = {
        'FromName': 'GEAD APP',
        'FromEmail': 'anthony.delpozo.m@gmail.com',
        'Subject': "Recuperar contraseña",
        'Text-Part': message,
        'Recipients': [{'Email': user.email}]
    }
    mailjet.send.create(email)


def send_notification(body, solicitude, topic):
    message = messaging.Message(
        data={
            "title":"GEAD APP",
            "body":body,
            "solicitude_id": str(solicitude.pk),
            "image_url": solicitude.image_url,
        },
        topic= topic,
    )
    response = messaging.send(message)
    print("RESPONSE", response)
    Notification(to="PFO", message="", theme="Nueva Solicitud", solicitude=solicitude).save()

@csrf_exempt
def send_massive_notification(request):
    body = request.POST["message"]
    print(body)
    if "authorities" in request.POST:
        print("Sending notification to AUTHORITIES")
        message = messaging.Message(
            data={
                "title":"Mensaje masivo",
                "body":body
            },
            topic= "AUT",
        )
        response = messaging.send(message)
        print("RESPONSE", response)

        Notification(to="AUT", message=body, theme="Mensaje", solicitude=None).save()


    if "focales" in request.POST:
        print("Sending notification to FOCALES")
        message = messaging.Message(
            data={
                "title":"Mensaje masivo",
                "body":body
            },
            topic= "PFO",
        )
        response = messaging.send(message)
        print("RESPONSE", response)

        Notification(to="PFO", message=body, theme="Mensaje", solicitude=None).save()


    # if "authorities" in request.POST and "focales" in request.POST:
    #     print("NOTIFICACION CREADA")
    #     Notification(to="ALL", message=body, theme="Mensaje", solicitude=None).save()

    solicitudes = Solicitude.objects.all().order_by('-date')
    args = {"solicitudes": solicitudes, "message": "Notificación enviada correctamente!"}
    return render(request,'user/home.html', args)





#-------------------------------WEB-------------------------------

def index(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        if username == "aplicativogead" and password == "Aplicativogead2019":
            print("ADENTRO")
            request.session['logged'] = True
            return redirect('/users/home')
        else:
            return render(request,'user/login.html', {"message": "Usuario o contraseña incorrecta"})
    elif request.method == 'GET':
        if request.session.has_key('logged'):
            if request.session['logged']:
                return redirect('/users/home')
    return render(request,'user/login.html')


def logout(request):
    if request.method == 'GET':
        if request.session.has_key('logged'):
            del request.session['logged']
            return redirect('/users/index')
        return redirect('/users/index')



def reset_password(request):
    print("reset_password")
    if request.method == 'GET':
        try:
            jwt.decode(request.GET['token'], settings.JWT_SECRET_KEY)['email']
            return render(request, 'user/reset_password.html')
        except Exception as e:
            return redirect('users:index')
    elif request.method == 'POST':
        data = request.POST.dict()
        token = request.GET['token']
        email = jwt.decode(token, settings.JWT_SECRET_KEY)['email']

        try:
            agent = Agent.objects.get(email=email)
            agent.password = pbkdf2_sha256.encrypt(data['password'],rounds=12000,salt_size=32)
            agent.save()
        except Agent.DoesNotExist as e:
            print(e)

        try:
            focal = EmpresaFocal.objects.get(email=email)
            focal.password = pbkdf2_sha256.encrypt(data['password'],rounds=12000,salt_size=32)
            focal.save()
        except EmpresaFocal.DoesNotExist as e:
            print(e)
        args = {"message": "Tu contraseña ha sido actualizada"}
        return render(request, 'user/reset_password.html', args )
    else:
        return HttpResponse(status= 405)


    # if request.method == 'POST':
    #     username = request.POST['username']
    #     password = request.POST['password']
    #     if username == "admin" and password == "admin":
    #         return redirect('/users/home')
    #     else:
    #         return render(request,'user/login.html', {"message": "Usuario o contraseña incorrecta"})
    # return render(request,'user/login.html')

def home(request):
    # return JSONResponse(CodeAccount.objects.all().values_list("code", "image"))
    print("HOME")
    if request.method == 'POST':
        return redirect('/users/home')
    elif request.method == 'GET':
        if request.session.has_key('logged'):
            logged = request.session['logged']
            if logged:
                solicitudes = Solicitude.objects.all().order_by('-date')
                args = {"solicitudes": solicitudes}
                return render(request,'user/home.html', args)
            return redirect('/users/index')
        return redirect('/users/index')

def detail(request, pk_solicitude):
    if request.method == 'POST':
        solicitude = Solicitude.objects.get(pk=pk_solicitude)

        if "select" in request.POST:
            solicitude.priority = request.POST['select']
        else:
            solicitude = Solicitude.objects.get(pk=pk_solicitude)
            args = {"solicitude": solicitude, "pk_solicitude": pk_solicitude, "message": True}
            return render(request,'user/detail.html', args)

        if "accepted" in request.POST and solicitude.accepted == False:
            solicitude.accepted = request.POST['accepted']
            send_email_accepted(solicitude.agent, solicitude)
            send_notification("Nueva solicitud {} creada".format(solicitude.title), solicitude, "PFO")
            print("SOLICITUD ACCEPTED")
        elif not "accepted" in request.POST:
            solicitude.accepted = False

        if "imgaccepted" in request.POST:
            solicitude.image_accepted = request.POST['imgaccepted']
        else:
            solicitude.image_accepted = False

        solicitude.receiver_name = request.POST['name']
        solicitude.receiver_dni = request.POST['dni']
        solicitude.receiver_phone = request.POST['phone']

        solicitude.save()

        return redirect('/users/detail/' + str(pk_solicitude))

    solicitude = Solicitude.objects.get(pk=pk_solicitude)
    helps = []
    for item in solicitude.item_set.all():
        helps.extend([ help for help in item.help_set.all()])
    args = {"solicitude": solicitude, "pk_solicitude": pk_solicitude, "helps": helps}
    return render(request,'user/detail.html', args)
